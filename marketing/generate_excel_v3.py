#!/usr/bin/env python3
"""Generate Excel from the full content plan."""
import re
import os

try:
    import openpyxl
except ImportError:
    os.system('pip install openpyxl')
    import openpyxl

def parse_content_plan(filename):
    """Parse the content plan markdown into structured data."""
    with open(filename, 'r', encoding='utf-8') as f:
        text = f.read()
    
    days = []
    current_day = None
    current_post = None
    
    lines = text.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Day header: ## 01.06 (ПН) — ...
        m = re.match(r'^##\s+(\d{2}\.\d{2})\s+\((\w+)\)\s*[—\-]\s*(.+)$', line)
        if m:
            if current_day:
                if current_post:
                    current_day['posts'].append(current_post)
                days.append(current_day)
            current_day = {
                'date': m.group(1),
                'day': m.group(2),
                'title': m.group(3),
                'posts': []
            }
            current_post = None
            i += 1
            continue
        
        # Post header: ### 🔹 Reels ... or ### 🔹 Текстовый пост ...
        m2 = re.match(r'^###\s+🔹\s+(.+?)\s+[\(（](.+?)[\)）]\s*[—\-–]\s*(.+)$', line)
        if not m2:
            m2 = re.match(r'^###\s+🔹\s+(.+?)\s*[—\-–]\s*(.+)$', line)
        if m2 and current_day:
            if current_post:
                current_day['posts'].append(current_post)
            groups = m2.groups()
            if len(groups) == 3:
                post_type = groups[0].strip()
                time_info = groups[1].strip()
                post_title = groups[2].strip()
            else:
                post_type = groups[0].strip()
                time_info = ''
                post_title = groups[1].strip()
            
            current_post = {
                'type': post_type,
                'time': time_info,
                'title': post_title,
                'doctor': '',
                'duration': '',
                'format': '',
                'scenario': '',
                'caption': ''
            }
            i += 1
            continue
        
        # Parse post details
        if current_post:
            # Doctor
            dm = re.match(r'^\*\*Врач:\*\*\s*(.+)$', line)
            if dm:
                current_post['doctor'] = dm.group(1)
                i += 1
                continue
            
            # Duration
            dm2 = re.match(r'^\*\*Длительность:\*\*\s*(.+)$', line)
            if dm2:
                current_post['duration'] = dm2.group(1)
                i += 1
                continue
            
            # Format
            dm3 = re.match(r'^\*\*Формат:\*\*\s*(.+)$', line)
            if dm3:
                current_post['format'] = dm3.group(1)
                i += 1
                continue
            
            # Scenario block
            if line == '**Сценарий:**' or line == '**Сценарий слайдов:**':
                j = i + 1
                scenario_lines = []
                while j < len(lines) and not lines[j].strip().startswith('**Текст для подписи:**') and not lines[j].strip().startswith('---'):
                    scenario_lines.append(lines[j].rstrip())
                    j += 1
                current_post['scenario'] = '\n'.join(scenario_lines).strip()
                i = j
                continue
            
            # Caption block
            if line == '**Текст для подписи:**':
                j = i + 1
                caption_lines = []
                in_code = False
                while j < len(lines):
                    l = lines[j].strip()
                    if l == '```':
                        in_code = not in_code
                        j += 1
                        continue
                    if not in_code and (l.startswith('---') or l.startswith('##') or l.startswith('###')):
                        break
                    if in_code or l:
                        caption_lines.append(lines[j].rstrip())
                    j += 1
                current_post['caption'] = '\n'.join(caption_lines).strip()
                i = j
                continue
        
        i += 1
    
    # Save last day
    if current_day:
        if current_post:
            current_day['posts'].append(current_post)
        days.append(current_day)
    
    return days

def generate_excel(days, output_file):
    """Generate Excel file from parsed data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Контент-план Июнь 2026"
    
    # Headers
    headers = ['Дата', 'День', 'Тема дня', 'Тип поста', 'Время',
               'Название', 'Врач', 'Длительность', 'Формат',
               'Сценарий', 'Текст для подписи']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        cell.fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    row = 2
    for day in days:
        if not day['posts']:
            # Day with no posts
            ws.cell(row=row, column=1, value=day['date'])
            ws.cell(row=row, column=2, value=day['day'])
            ws.cell(row=row, column=3, value=day['title'])
            row += 1
        else:
            for post in day['posts']:
                ws.cell(row=row, column=1, value=day['date'])
                ws.cell(row=row, column=2, value=day['day'])
                ws.cell(row=row, column=3, value=day['title'])
                ws.cell(row=row, column=4, value=post['type'])
                ws.cell(row=row, column=5, value=post['time'])
                ws.cell(row=row, column=6, value=post['title'])
                ws.cell(row=row, column=7, value=post['doctor'])
                ws.cell(row=row, column=8, value=post['duration'])
                ws.cell(row=row, column=9, value=post['format'])
                ws.cell(row=row, column=10, value=post['scenario'])
                ws.cell(row=row, column=11, value=post['caption'])
                row += 1
    
    # Column widths
    widths = [10, 8, 25, 15, 8, 35, 20, 12, 20, 50, 50]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Text wrap for scenario and caption
    for r in range(2, row):
        for c in [10, 11]:
            cell = ws.cell(row=r, column=c)
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
    
    wb.save(output_file)
    print(f'Excel saved to {output_file}')
    print(f'Total rows: {row - 2}')

if __name__ == '__main__':
    days = parse_content_plan('content_plan_june_2026.md')
    print(f'Parsed {len(days)} days')
    generate_excel(days, 'content_plan_june_2026.xlsx')
