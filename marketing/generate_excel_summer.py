#!/usr/bin/env python3
"""Generate Excel from the summer content plan."""
import re
import os

try:
    import openpyxl
except ImportError:
    os.system('pip install openpyxl')
    import openpyxl

def parse_content_plan(filename):
    """Parse the content plan markdown into structured data."""
    with open(filename, 'r', encoding='utf-8') as f:
        text = f.read()
    
    days = []
    current_day = None
    current_post = None
    
    lines = text.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Day header: ### 16.06 (ВТ) — ...
        m = re.match(r'^###\s+(\d{2}\.\d{2})\s+\((\w+)\)\s*[—\-–]\s*(.+)$', line)
        if m:
            if current_day:
                if current_post:
                    current_day['posts'].append(current_post)
                days.append(current_day)
            current_day = {
                'date': m.group(1),
                'day': m.group(2),
                'title': m.group(3),
                'posts': []
            }
            current_post = None
            i += 1
            continue
        
        # Post type: **Reels (10:00)** or **Пост-карусель (18:00)** or **Пост до/после (18:00)**
        m2 = re.match(r'^\*\*(.+?)\s*\((\d+:\d+)\)\*\*', line)
        if m2 and current_day:
            if current_post:
                current_day['posts'].append(current_post)
            current_post = {
                'type': m2.group(1).strip(),
                'time': m2.group(2).strip(),
                'hook': '',
                'doctor_text': '',
                'slides': '',
                'caption': '',
                'photo_desc': ''
            }
            i += 1
            continue
        
        if current_post:
            # Hook
            hm = re.match(r'^Хук:\s*[«\"](.+)[»\"]', line)
            if hm:
                current_post['hook'] = hm.group(1)
                i += 1
                continue
            
            # Doctor text
            dm = re.match(r'^Текст врача\s*\([\d\-\sсек]+\):\s*[«\"](.+)[»\"]', line)
            if dm:
                current_post['doctor_text'] = dm.group(1)
                i += 1
                continue
            
            # Photo description
            pm = re.match(r'^Описание:\s*[«\"](.+)[»\"]', line)
            if pm:
                current_post['photo_desc'] = pm.group(1)
                i += 1
                continue
            
            # Тема for carousel
            tm = re.match(r'^Тема:\s*[«\"](.+)[»\"]', line)
            if tm:
                current_post['hook'] = tm.group(1)
                i += 1
                continue
            
            # Slides
            sm = re.match(r'^Слайд\s+\d+:\s*(.+)$', line)
            if sm:
                current_post['slides'] += sm.group(1) + '\n'
                i += 1
                continue
            
            # Caption
            cm = re.match(r'^Текст для поста:\s*[«\"](.+)[»\"]', line)
            if cm:
                current_post['caption'] = cm.group(1)
                i += 1
                continue
            
            # Multi-line caption
            if line.startswith('«') and current_post['caption']:
                current_post['caption'] += ' ' + line.strip('«»')
            elif line.endswith('»') and current_post['caption']:
                current_post['caption'] += ' ' + line.strip('«»')
        
        i += 1
    
    if current_day:
        if current_post:
            current_day['posts'].append(current_post)
        days.append(current_day)
    
    return days

def generate_excel(days, output_file):
    """Generate Excel file from parsed data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Контент-план ARclinic"
    
    # Headers
    headers = ['Дата', 'День', 'Тема дня', 'Тип поста', 'Время',
               'Хук / Тема', 'Текст врача / Слайды', 'Текст для поста', 'Описание фото']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        cell.fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    row = 2
    for day in days:
        if not day['posts']:
            ws.cell(row=row, column=1, value=day['date'])
            ws.cell(row=row, column=2, value=day['day'])
            ws.cell(row=row, column=3, value=day['title'])
            row += 1
        else:
            for post in day['posts']:
                ws.cell(row=row, column=1, value=day['date'])
                ws.cell(row=row, column=2, value=day['day'])
                ws.cell(row=row, column=3, value=day['title'])
                ws.cell(row=row, column=4, value=post['type'])
                ws.cell(row=row, column=5, value=post['time'])
                ws.cell(row=row, column=6, value=post['hook'])
                
                doctor_or_slides = post['doctor_text'] if post['doctor_text'] else post['slides']
                ws.cell(row=row, column=7, value=doctor_or_slides)
                ws.cell(row=row, column=8, value=post['caption'])
                ws.cell(row=row, column=9, value=post['photo_desc'])
                row += 1
    
    # Column widths
    widths = [10, 8, 30, 15, 8, 40, 60, 60, 40]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Text wrap
    for r in range(2, row):
        for c in [6, 7, 8, 9]:
            cell = ws.cell(row=r, column=c)
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
    
    wb.save(output_file)
    print(f'Excel saved to {output_file}')
    print(f'Total rows: {row - 2}')

if __name__ == '__main__':
    days = parse_content_plan('content_plan_summer_2026.md')
    print(f'Parsed {len(days)} days')
    for d in days:
        print(f"  {d['date']} ({d['day']}) - {d['title']}: {len(d['posts'])} posts")
    generate_excel(days, 'content_plan_summer_2026.xlsx')
