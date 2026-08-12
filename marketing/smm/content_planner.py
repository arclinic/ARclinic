"""
Генератор контент-планов для соцсетей ARclinic.
Учитывает рубрики, сезонность, загрузку врачей, акции.
"""

import os
import json
from datetime import datetime, timedelta
from typing import Optional
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent
BRAIN = ROOT / "мозг_клиники_ARclinic"


class ContentPlanner:
    """Планировщик контента для Instagram, Telegram, VK."""

    RUBRICS = {
        "education": {"name": "Образование", "goal": "Формировать доверие через экспертизу"},
        "cases": {"name": "Кейсы до/после", "goal": "Визуальный результат, доверие"},
        "lifestyle": {"name": "Лайфстайл врачей", "goal": "Показать людей за халатами"},
        "behind_scenes": {"name": "Закулисье клиники", "goal": "Показать стандарты и процессы"},
        "qa": {"name": "Ответы на вопросы", "goal": "Закрывать возражения, повышать охваты"},
        "anna_brand": {"name": "Личный бренд Анны", "goal": "Ассоциировать клинику с основателем"},
    }

    SEASONS = {
        (1, 2): {
            "name": "Январь–февраль",
            "themes": [
                "Восстановление после праздников",
                "Пилинги (зима — сезон)",
                "Лазерное омоложение",
                "Подготовка к весне",
            ],
        },
        (3, 5): {
            "name": "Март–май (пик)",
            "themes": [
                "Ботулинотерапия перед летом",
                "Контурная пластика",
                "Плазмотерапия",
                "Чекапы «здоровье перед отпуском»",
            ],
        },
        (6, 8): {
            "name": "Июнь–август (спад)",
            "themes": [
                "Уходовая косметология",
                "Консультации, планирование на осень",
                "Образовательный контент",
                "Истории врачей, лайфстайл",
            ],
        },
        (9, 11): {
            "name": "Сентябрь–ноябрь (пик)",
            "themes": [
                "Восстановление после лета",
                "Ботулинотерапия, филлеры, нити",
                "Чекапы «здоровье на осень»",
                "Подготовка к зиме",
            ],
        },
        (12, 12): {
            "name": "Декабрь",
            "themes": [
                "Процедуры «к Новому году»",
                "Подарочные сертификаты",
                "Итоги года клиники",
            ],
        },
    }

    DOCTORS_BY_SPEC = {
        "косметология": ["Ряховская", "Батиенко", "Фокина", "Подольникова", "Зорина"],
        "гинекология": ["Афанасьева"],
        "неврология": ["Рубаник"],
        "дерматология": ["Зорина", "Дроздова"],
        "эндокринология": ["Львович"],
        "антивозрастная медицина": ["Резник Анна"],
        "инъекции": ["Ряховская", "Батиенко", "Фокина"],
        "лазер": ["Подольникова", "Батиенко"],
        "рф-лифтинг": ["Ряховская", "Труфанов"],
    }

    def __init__(self, month: Optional[str] = None):
        if month:
            self.year, self.month_num = map(int, month.split("-"))
        else:
            now = datetime.now()
            self.year = now.year
            self.month_num = now.month
        self._load_context()

    def _load_context(self):
        """Загружает контекст из мозга клиники."""
        self.tone = ""
        self.principles = ""
        self.doctors = ""
        tone_path = BRAIN / "голос" / "tone.md"
        principles_path = BRAIN / "идентичность" / "контент-принципы.md"
        doctors_path = BRAIN / "бизнес" / "клиника.md"
        if tone_path.exists():
            self.tone = tone_path.read_text(encoding="utf-8")
        if principles_path.exists():
            self.principles = principles_path.read_text(encoding="utf-8")
        if doctors_path.exists():
            self.doctors = doctors_path.read_text(encoding="utf-8")

    def get_season_themes(self) -> list:
        """Возвращает сезонные темы для текущего месяца."""
        for (start, end), data in self.SEASONS.items():
            if start <= self.month_num <= end:
                return data["themes"]
        return []

    def get_rubric_rotation(self) -> list:
        """Ротация рубрик на неделю (6 дней, ВС — выходной)."""
        return [
            ("education", "Reels"),
            ("cases", "Пост"),
            ("lifestyle", "Stories"),
            ("qa", "Reels"),
            ("behind_scenes", "Пост"),
            ("anna_brand", "Reels"),
        ]

    def generate_weekly_plan(self, start_date: datetime, doctors: Optional[list] = None) -> str:
        """Генерирует контент-план на неделю."""
        rubric_rotation = self.get_rubric_rotation()
        season_themes = self.get_season_themes()
        lines = []

        dates = [(start_date + timedelta(days=i)).strftime("%d.%m") for i in range(6)]
        day_names_ru = ["ПН", "ВТ", "СР", "ЧТ", "ПТ", "СБ"]

        lines.append(f"# Контент-план — Неделя {start_date.isocalendar()[1]} ({start_date.strftime('%d.%m')}–{(start_date + timedelta(days=5)).strftime('%d.%m')})\n")

        for i, (date, day) in enumerate(zip(dates, day_names_ru)):
            rubric_key, fmt = rubric_rotation[i]
            rubric = self.RUBRICS[rubric_key]
            season_hint = season_themes[i % len(season_themes)] if season_themes else ""

            lines.append(f"## {date} ({day}) — {rubric['name']}")
            lines.append(f"**Формат:** {fmt}")
            lines.append(f"**Рубрика:** {rubric['name']} | {rubric['goal']}")
            if season_hint:
                lines.append(f"**Сезонная тема:** {season_hint}")
            lines.append("**Тема:** [сгенерировать]\n")

        return "\n".join(lines)

    def generate_monthly_plan(self, doctors: Optional[list] = None) -> str:
        """Генерирует контент-план на месяц."""
        first_day = datetime(self.year, self.month_num, 1)
        season_themes = self.get_season_themes()
        lines = []

        months_ru = [
            "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
            "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
        ]
        month_name = months_ru[self.month_num]

        lines.append(f"# Контент-план Instagram — {month_name} {self.year}")
        lines.append("## Клиника: ARclinic — Центр антивозрастной медицины и косметологии в Санкт-Петербурге")
        lines.append("## Формат: Reels + текстовый пост ежедневно\n")

        lines.append("## Сезонные темы месяца:")
        for theme in season_themes:
            lines.append(f"- {theme}")
        lines.append("")

        lines.append("## 📌 Рубрики месяца (ротация по дням):")
        lines.append("| День | Формат | Рубрика |")
        lines.append("|------|--------|---------|")
        rubric_rotation = self.get_rubric_rotation()
        day_names_ru = ["ПН", "ВТ", "СР", "ЧТ", "ПТ", "СБ"]
        for day, (rubric_key, fmt) in zip(day_names_ru, rubric_rotation):
            lines.append(f"| {day} | {fmt} | {self.RUBRICS[rubric_key]['name']} |")
        lines.append("")

        weeks_in_month = 0
        current = first_day
        while current.month == self.month_num:
            if current.weekday() < 6:
                weeks_in_month += 1
                lines.append(self.generate_weekly_plan(current, doctors))
                lines.append("")
                current += timedelta(days=6 - current.weekday())
            current += timedelta(days=1)

        return "\n".join(lines)

    def export_to_xlsx(self, content: str, output_path: str):
        """Экспортирует контент-план в Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "Контент-план"

            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            headers = ["Дата", "День", "Формат", "Рубрика", "Тема", "Врач", "Статус"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            row = 2
            for line in content.split("\n"):
                if line.startswith("## ") and ("(" in line):
                    parts = line.replace("## ", "").split(" — ")
                    if len(parts) >= 2:
                        date_part = parts[0].strip()
                        rubric = parts[1].strip()
                        ws.cell(row=row, column=1, value=date_part).border = thin_border
                        ws.cell(row=row, column=3, value="Reels").border = thin_border
                        ws.cell(row=row, column=4, value=rubric).border = thin_border
                        if "Reels" in rubric or "reels" in rubric.lower():
                            for c in range(1, 8):
                                ws.cell(row=row, column=c).fill = green_fill
                        row += 1

            for col in range(1, 8):
                ws.column_dimensions[chr(64 + col)].width = 18 if col <= 4 else 25
            wb.save(output_path)
        except ImportError:
            with open(output_path.replace(".xlsx", ".md"), "w", encoding="utf-8") as f:
                f.write(content)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Генератор контент-планов ARclinic")
    parser.add_argument("--month", help="Месяц в формате YYYY-MM", default=None)
    parser.add_argument("--doctors", help="Врачи через запятую", default="")
    parser.add_argument("--output", help="Путь для сохранения", default="")
    parser.add_argument("--export", action="store_true", help="Экспорт в XLSX")
    parser.add_argument("--style", choices=["minimal", "full"], default="minimal",
                        help="Стиль плана: minimal (сетка) или full (с полными текстами)")

    args = parser.parse_args()
    planner = ContentPlanner(month=args.month or None)
    doctors = [d.strip() for d in args.doctors.split(",") if d.strip()] if args.doctors else None

    if args.style == "minimal":
        content = planner.generate_monthly_plan(doctors)
    else:
        print("Full style: используйте generate_content_plan.py для генерации полных текстов")
        content = planner.generate_monthly_plan(doctors)

    output = args.output or f"content_plan_{planner.month_num:02d}_{planner.year}.md"
    if args.export:
        planner.export_to_xlsx(content, output.replace(".md", ".xlsx"))
    else:
        with open(output, "w", encoding="utf-8") as f:
            f.write(content)
    print(f"План сохранён в {output}")
