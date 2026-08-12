"""
Аналитика соцсетей ARclinic.
Сбор метрик: Telegram, VK, YouTube, Rutube, Дзен.
Сводный отчёт в XLSX.
"""

import os
import sys
import json
import requests
from datetime import datetime, timedelta
from typing import Optional, Dict, Any
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))

from channels import CHANNELS, get_channel_id, list_channels


class SocialAnalytics:
    """Сборщик аналитики по соцсетям ARclinic."""

    def __init__(self):
        self.tg_bot_token = os.getenv("TG_BOT_TOKEN", "")
        self.tg_channel = os.getenv("TG_CHANNEL_ID", "@arclinic1")
        self.vk_token = os.getenv("VK_ACCESS_TOKEN", "")
        self.vk_group_id = os.getenv("VK_GROUP_ID", "")
        self.youtube_api_key = os.getenv("YOUTUBE_API_KEY", "")
        self.reports_dir = ROOT / "reports"

    def get_telegram_stats(self) -> dict:
        """Метрики Telegram-канала."""
        stats = {
            "platform": "Telegram",
            "subscribers": "ручной сбор / TGStat",
            "avg_views": "ручной сбор",
            "avg_engagement": "ручной сбор",
            "posts_this_week": 0,
            "note": "Telegram Bot API не отдаёт статистику канала. Используйте TGStat API (платный) или ручной сбор.",
        }
        return stats

    def get_vk_stats(self) -> dict:
        """Метрики VK-сообщества через VK API."""
        if not self.vk_token or not self.vk_group_id:
            return {"platform": "VK", "error": "Нет токена или group_id в .env"}

        stats = {"platform": "VK", "subscribers": 0, "posts": [], "error": None}
        try:
            url = "https://api.vk.com/method/groups.getMembers"
            params = {"group_id": self.vk_group_id, "v": "5.199", "access_token": self.vk_token}
            resp = requests.get(url, params=params, timeout=15).json()
            if "response" in resp:
                stats["subscribers"] = resp["response"].get("count", 0)

            wall_url = "https://api.vk.com/method/wall.get"
            wall_params = {
                "owner_id": f"-{self.vk_group_id}", "count": 10,
                "v": "5.199", "access_token": self.vk_token,
            }
            wall_resp = requests.get(wall_url, params=wall_params, timeout=15).json()
            if "response" in wall_resp:
                for item in wall_resp["response"]["items"]:
                    stats["posts"].append({
                        "id": item["id"],
                        "date": datetime.fromtimestamp(item["date"]).isoformat(),
                        "likes": item.get("likes", {}).get("count", 0),
                        "comments": item.get("comments", {}).get("count", 0),
                        "reposts": item.get("reposts", {}).get("count", 0),
                        "views": item.get("views", {}).get("count", 0),
                    })
        except Exception as e:
            stats["error"] = str(e)
        return stats

    def get_youtube_stats(self) -> dict:
        """Метрики YouTube-канала."""
        if not self.youtube_api_key:
            return {"platform": "YouTube", "error": "Нет YOUTUBE_API_KEY в .env"}

        stats = {"platform": "YouTube", "subscribers": 0, "videos": [], "error": None}
        try:
            from googleapiclient.discovery import build
            youtube = build("youtube", "v3", developerKey=self.youtube_api_key)

            ch_resp = youtube.channels().list(part="statistics", forHandle="@arclinic").execute()
            if ch_resp.get("items"):
                s = ch_resp["items"][0]["statistics"]
                stats["subscribers"] = int(s.get("subscriberCount", 0))
                stats["total_views"] = int(s.get("viewCount", 0))
                stats["total_videos"] = int(s.get("videoCount", 0))
        except ImportError:
            stats["error"] = "google-api-python-client не установлен"
        except Exception as e:
            stats["error"] = str(e)
        return stats

    def generate_report(self, period: str = "week") -> str:
        """Генерирует сводный отчёт по всем платформам."""
        platforms = {
            "telegram": self.get_telegram_stats(),
            "vk": self.get_vk_stats(),
            "youtube": self.get_youtube_stats(),
        }

        date_str = datetime.now().strftime("%Y-%m-%d")
        lines = [
            f"# Аналитика соцсетей ARclinic — {date_str}",
            f"**Период:** {period}",
            "",
        ]

        for name, stats in platforms.items():
            lines.append(f"## {stats.get('platform', name)}")
            if stats.get("error"):
                lines.append(f"⚠️ Ошибка: {stats['error']}")
            else:
                for key, value in stats.items():
                    if key not in ("platform", "error", "posts"):
                        lines.append(f"- **{key}:** {value}")
            lines.append("")

        return "\n".join(lines)

    def export_to_xlsx(self, report: str, output_path: str):
        """Экспортирует отчёт в XLSX."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "Аналитика соцсетей"

            headers = ["Платформа", "Метрика", "Значение", "Примечание"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")

            row = 2
            current_platform = ""
            for line in report.split("\n"):
                if line.startswith("## "):
                    current_platform = line.replace("## ", "")
                elif line.startswith("- **"):
                    parts = line.replace("- **", "").split(":** ", 1)
                    if len(parts) == 2:
                        ws.cell(row=row, column=1, value=current_platform)
                        ws.cell(row=row, column=2, value=parts[0])
                        ws.cell(row=row, column=3, value=parts[1])
                        row += 1

            for col in range(1, 5):
                ws.column_dimensions[chr(64 + col)].width = 25
            wb.save(output_path)
        except ImportError:
            with open(output_path.replace(".xlsx", ".md"), "w", encoding="utf-8") as f:
                f.write(report)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Аналитика соцсетей ARclinic")
    parser.add_argument("--period", choices=["day", "week", "month"], default="week",
                        help="Период анализа")
    parser.add_argument("--export", action="store_true", help="Экспорт в XLSX")

    args = parser.parse_args()
    analytics = SocialAnalytics()
    report = analytics.generate_report(args.period)
    print(report)

    if args.export:
        date_str = datetime.now().strftime("%Y-%m-%d")
        path = str(ROOT / "reports" / f"social_analytics_{date_str}.xlsx")
        analytics.export_to_xlsx(report, path)
        print(f"\nОтчёт сохранён в {path}")
