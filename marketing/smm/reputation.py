"""
Мониторинг репутации ARclinic.
Сбор отзывов: Yandex Maps, 2GIS, Google Maps, ProDoctors.
Анализ тональности, тренды, алерты.
"""

import os
import json
import requests
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List
from pathlib import Path
from collections import Counter

ROOT = Path(__file__).resolve().parent.parent.parent
BRAIN = ROOT / "мозг_клиники_ARclinic"


class ReputationMonitor:
    """Мониторинг репутации клиники на внешних площадках."""

    PLATFORMS = {
        "yandex_maps": {"name": "Яндекс Карты", "url": "https://yandex.ru/maps/org/arclinic/"},
        "2gis": {"name": "2GIS", "url": "https://2gis.ru/spb/firm/..."},
        "google_maps": {"name": "Google Maps", "url": "https://maps.google.com/..."},
        "prodoctorov": {"name": "ПроДокторов", "url": "https://prodoctorov.ru/spb/lpu/..."},
    }

    def __init__(self):
        self.reviews_dir = ROOT / "reports" / "reviews"
        self.reviews_dir.mkdir(parents=True, exist_ok=True)
        self._load_existing()

    def _load_existing(self):
        """Загружает историю отзывов из мозга клиники."""
        reviews_path = BRAIN / "бизнес" / "отзывы.md"
        self.existing_reviews = reviews_path.read_text(encoding="utf-8") if reviews_path.exists() else ""

    def check_reviews(self) -> dict:
        """Проверяет наличие новых отзывов на всех площадках."""
        results = {}
        for key, platform in self.PLATFORMS.items():
            results[key] = {
                "name": platform["name"],
                "url": platform["url"],
                "status": "manual_check",
                "note": "Требуется ручной сбор или парсинг",
                "rating": "см. отзывы.md",
            }
        return results

    def analyze_sentiment(self, reviews: List[str]) -> dict:
        """Анализирует тональность отзывов."""
        total = len(reviews)
        if total == 0:
            return {"total": 0, "positive": 0, "neutral": 0, "negative": 0, "score": 0}

        pos_words = {"спасибо", "благодар", "отлично", "понравил", "рекоменд", "профессионал",
                     "чист", "уют", "вниматель", "результат", "эффект", "качество", "лучш",
                     "комфорт", "приятн", "вежлив", "грамотн", "прекрасн", "супер", "довол",
                     "уютн", "красив", "аккуратн", "забот", "тепл", "душевн", "понрав"}
        neg_words = {"ужас", "кошмар", "отвратительн", "обман", "навязыва", "больно",
                     "дорого", "плох", "разочарован", "недовол", "ужасн", "проблем",
                     "осложнен", "шрам", "груб", "хам", "грязн", "неприятн"}

        positive = 0
        negative = 0
        neutral = 0

        for review in reviews:
            review_lower = review.lower()
            pos_count = sum(1 for w in pos_words if w in review_lower)
            neg_count = sum(1 for w in neg_words if w in review_lower)
            if pos_count > neg_count:
                positive += 1
            elif neg_count > pos_count:
                negative += 1
            else:
                neutral += 1

        score = (positive / total - negative / total) * 100 if total > 0 else 0
        return {
            "total": total, "positive": positive, "neutral": neutral,
            "negative": negative, "score": round(score, 1),
        }

    def generate_report(self) -> str:
        """Генерирует отчёт по репутации."""
        date_str = datetime.now().strftime("%Y-%m-%d")
        platform_status = self.check_reviews()

        lines = [
            f"# Мониторинг репутации ARclinic — {date_str}",
            "",
            "## Площадки и статус",
            "| Площадка | Рейтинг | Статус |",
            "|----------|---------|--------|",
        ]
        for key, info in platform_status.items():
            lines.append(f"| {info['name']} | {info['rating']} | {info['status']} |")

        lines.extend([
            "",
            "## Анализ тональности",
            "[требуется загрузка актуальных отзывов]",
            "",
            "## Тренды",
            "- Рейтинг за месяц: [собрать]",
            "- Динамика: [сравнить с прошлым месяцем]",
            "- Топ-3 зоны позитива: [определить]",
            "- Топ-3 зоны негатива: [определить]",
            "",
            "## Алерты",
            "Новых негативных отзывов: [проверить]",
            "Ответы требуются: [проверить]",
        ])

        return "\n".join(lines)

    def export_to_xlsx(self, report: str, output_path: str):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "Репутация"

            headers = ["Площадка", "Рейтинг", "Отзывов всего", "Позитив", "Нейтрально",
                       "Негатив", "Тренд", "Дата проверки"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")

            for col in range(1, 9):
                ws.column_dimensions[chr(64 + col)].width = 18
            wb.save(output_path)
        except ImportError:
            pass


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Мониторинг репутации ARclinic")
    parser.add_argument("--check", action="store_true", help="Проверить новые отзывы")
    parser.add_argument("--report", action="store_true", help="Сгенерировать отчёт")

    args = parser.parse_args()
    monitor = ReputationMonitor()

    if args.check:
        results = monitor.check_reviews()
        for key, info in results.items():
            print(f"{info['name']}: {info['status']} ({info['url']})")
    elif args.report:
        report = monitor.generate_report()
        print(report)
    else:
        print("Укажите --check или --report")
