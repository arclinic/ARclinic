import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date
import os

def create_report(data, output_dir="reports"):
    os.makedirs(output_dir, exist_ok=True)
    filename = f"report-{date.today().isoformat()}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()

    # Sheet 1 - Summary
    ws1 = wb.active
    ws1.title = "Сводка"
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws1.cell(row=1, column=1, value="Метрика").font = header_font_white
    ws1.cell(row=1, column=1).fill = header_fill
    ws1.cell(row=1, column=1).border = thin_border
    ws1.cell(row=1, column=2, value="Значение").font = header_font_white
    ws1.cell(row=1, column=2).fill = header_fill
    ws1.cell(row=1, column=2).border = thin_border

    metrics = [
        ("Дата", data.get("date", date.today().isoformat())),
        ("Выручка за месяц", data.get("revenue", "нет данных")),
        ("Выполнение плана", data.get("plan_progress", "нет данных")),
        ("Средний чек", data.get("avg_check", "нет данных")),
        ("Активные клиенты", data.get("active_clients", "нет данных")),
        ("Новые клиенты", data.get("new_clients", "нет данных")),
        ("Загрузка врачей", data.get("doctor_load", "нет данных")),
    ]

    for i, (metric, value) in enumerate(metrics, 2):
        ws1.cell(row=i, column=1, value=metric).border = thin_border
        ws1.cell(row=i, column=2, value=str(value)).border = thin_border

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20

    # Sheet 2 - Details
    ws2 = wb.create_sheet("Детали")
    ws2.cell(row=1, column=1, value="Категория").font = header_font_white
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=1).border = thin_border
    ws2.cell(row=1, column=2, value="Детали").font = header_font_white
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=2).border = thin_border

    details = [
        ("Флаги", data.get("flags", "нет")),
        ("Фазы из plans/", data.get("phases", "нет")),
        ("Git-изменения", data.get("git_changes", "нет")),
        ("Заметки", data.get("notes", "нет")),
    ]

    for i, (cat, val) in enumerate(details, 2):
        ws2.cell(row=i, column=1, value=cat).border = thin_border
        ws2.cell(row=i, column=2, value=str(val)).border = thin_border

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 60

    wb.save(filepath)
    return filepath
