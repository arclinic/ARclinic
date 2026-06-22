import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, datetime
import os

def create_task(task_data, output_dir="reports"):
    os.makedirs(output_dir, exist_ok=True)
    filename = f"tasks-{date.today().isoformat()}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws = wb.active
    ws.title = "Задачи"

    headers = ["ID", "Задача", "Ответственный", "Дедлайн", "Приоритет", "Статус"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    tasks = task_data.get("tasks", [])
    for i, t in enumerate(tasks, 2):
        ws.cell(row=i, column=1, value=t.get("id", i - 1)).border = thin_border
        ws.cell(row=i, column=2, value=t.get("title", "")).border = thin_border
        ws.cell(row=i, column=3, value=t.get("assignee", "")).border = thin_border
        ws.cell(row=i, column=4, value=t.get("deadline", "")).border = thin_border
        ws.cell(row=i, column=5, value=t.get("priority", "")).border = thin_border
        status = t.get("status", "pending")
        cell = ws.cell(row=i, column=6, value=status)
        cell.border = thin_border
        if status == "overdue":
            cell.fill = red_fill
        elif status == "done":
            cell.fill = green_fill

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15

    # Sheet 2 - Meeting protocols
    ws2 = wb.create_sheet("Протоколы")
    proto_headers = ["Дата", "Тема", "Участники", "Решения", "Ответственные", "Срок", "Статус"]
    for col, h in enumerate(proto_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    protocols = task_data.get("protocols", [])
    for i, p in enumerate(protocols, 2):
        ws2.cell(row=i, column=1, value=p.get("date", "")).border = thin_border
        ws2.cell(row=i, column=2, value=p.get("topic", "")).border = thin_border
        ws2.cell(row=i, column=3, value=p.get("participants", "")).border = thin_border
        ws2.cell(row=i, column=4, value=p.get("decisions", "")).border = thin_border
        ws2.cell(row=i, column=5, value=p.get("responsible", "")).border = thin_border
        ws2.cell(row=i, column=6, value=p.get("deadline", "")).border = thin_border
        ws2.cell(row=i, column=7, value=p.get("status", "open")).border = thin_border

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 45
    ws2.column_dimensions['E'].width = 25
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 12

    # Sheet 3 - Checklists
    ws3 = wb.create_sheet("Чеклисты")
    cl_headers = ["Чеклист", "Пункт", "Ответственный", "Статус", "Дата выполнения"]
    for col, h in enumerate(cl_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    checklists = task_data.get("checklists", [])
    for i, cl in enumerate(checklists, 2):
        ws3.cell(row=i, column=1, value=cl.get("list_name", "")).border = thin_border
        ws3.cell(row=i, column=2, value=cl.get("item", "")).border = thin_border
        ws3.cell(row=i, column=3, value=cl.get("assignee", "")).border = thin_border
        status = cl.get("status", "pending")
        cell = ws3.cell(row=i, column=4, value=status)
        cell.border = thin_border
        if status == "done":
            cell.fill = green_fill
        ws3.cell(row=i, column=5, value=cl.get("completed_at", "")).border = thin_border

    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 15

    wb.save(filepath)
    return filepath


def create_checklist(checklist_data, output_dir="reports"):
    return create_task({"tasks": [], "protocols": [], "checklists": checklist_data}, output_dir)


def create_protocol(protocol_data, output_dir="reports"):
    return create_task({"tasks": [], "protocols": protocol_data, "checklists": []}, output_dir)
