#!/usr/bin/env python3
"""
ARclinic Doctor Retention Agent
Возвращаемость пациентов к врачам: расчёт, рейтинг, тренды, алерты, прогноз.

Режимы:
  monthly — полный XLSX-отчёт + Markdown-сводка
  weekly  — быстрая проверка Rate A за 3 мес + алерты
"""

import sys
import os
import json
import calendar
import argparse
from datetime import date, datetime, timedelta
from collections import defaultdict
from pathlib import Path
from typing import Any

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))))
sys.path.insert(0, ROOT)
from shared.bitrix24 import call

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# Constants
# =============================================================================

DOCTOR_MAP: dict[int, str] = {
    9821: "Резник Анна",
    5749: "Афанасьева Лилия Борисовна",
    10029: "Батиенко Дарья Дмитриевна",
    10031: "Львович Ирина Викторовна",
    10033: "Ряховская Наталья Дмитриевна",
    10035: "Дроздова Анна Андреевна",
    10037: "Фокина Екатерина Евгеньевна",
    7135: "Яхина Алиса Альфировна",
    10763: "Зорина Надежда Викторовна",
    15353: "Подольникова Мария Игоревна",
    21191: "Труфанов Георгий Сергеевич",
    10041: "Яковлева Ольга Михайловна",
    10043: "Рубаник Кирилл Сергеевич",
    20629: "Плотникова Анна Юрьевна",
}

EXCLUDED_IDS: set[int] = {13505, 10027, 11145}
ACTIVE_IDS: list[int] = [k for k in DOCTOR_MAP if k not in EXCLUDED_IDS]

WINDOWS: list[int] = [3, 6, 9, 12]
MIN_PRIMARY_PATIENTS = 10
CACHE_FILE = "retention_cache.json"
DATA_START = date(2025, 1, 1)

# XLSX styles
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# =============================================================================
# Helpers
# =============================================================================

def add_months(dt: date, months: int) -> date:
    """Прибавить N месяцев к дате с clamp'ом на конец месяца."""
    month = dt.month + months
    year = dt.year + (month - 1) // 12
    month = (month - 1) % 12 + 1
    max_day = calendar.monthrange(year, month)[1]
    return date(year, month, min(dt.day, max_day))


def pct_str(value: float | None) -> str:
    """Формат процента для вывода."""
    if value is None:
        return "—"
    return f"{value:.1f}%"


def format_rate(value: float) -> str:
    return f"{value * 100:.1f}%"


# =============================================================================
# Agent
# =============================================================================

class DoctorRetentionAgent:
    """Агент расчёта возвращаемости пациентов к врачам."""

    def __init__(self, report_date: date | None = None, output_dir: str = "reports"):
        self.report_date = report_date or date.today()
        self.output_dir = output_dir
        self.cache_path = os.path.join(output_dir, CACHE_FILE)
        self.deals: list[dict] = []
        self.visits: dict[int, dict[int, list[date]]] = {}  # doctor_id -> contact_id -> [visit dates]

    # -------------------------------------------------------------------------
    # Data fetch & cache
    # -------------------------------------------------------------------------

    def load_data(self, force_refresh: bool = False, silent: bool = False) -> None:
        """Загрузить сделки из кеша или API."""
        if not force_refresh and self._cache_valid():
            self._load_cache()
            if not silent:
                print(f"[cache] Загружено {len(self.deals)} сделок из кеша ({self.cache_path})")
            self._build_visits()
            return

        if not silent:
            print("[api] Запрашиваю сделки из Bitrix24...")
        try:
            self.deals = self._fetch_all_deals()
            if not silent:
                print(f"[api] Получено {len(self.deals)} сделок")
            self._save_cache()
        except Exception as e:
            print(f"[api] Ошибка при запросе к Bitrix24: {e}")
            # Пробуем загрузить кеш, даже если он устарел
            if os.path.exists(self.cache_path):
                print("[api] Использую устаревший кеш как fallback.")
                self._load_cache()
            else:
                print("[api] Кеш отсутствует, данных нет.")
                self.deals = []
        self._build_visits()

    def _cache_valid(self) -> bool:
        if not os.path.exists(self.cache_path):
            return False
        try:
            with open(self.cache_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            cached_date = data.get("fetch_date", "")
            return cached_date == self.report_date.isoformat()
        except Exception:
            return False

    def _load_cache(self) -> None:
        with open(self.cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        self.deals = data.get("deals", [])

    def _save_cache(self) -> None:
        os.makedirs(self.output_dir, exist_ok=True)
        with open(self.cache_path, "w", encoding="utf-8") as f:
            json.dump({
                "fetch_date": self.report_date.isoformat(),
                "deals": self.deals,
            }, f, ensure_ascii=False, default=str)

    def _fetch_all_deals(self) -> list[dict]:
        """Пагинированный сбор всех завершённых сделок с 2025-01-01."""
        all_deals: list[dict] = []
        start = 0
        while True:
            try:
                result = call("crm.deal.list", {
                    "filter": {
                        "STAGE_ID": "C1:WON",
                        ">=DATE_CREATE": DATA_START.isoformat(),
                    },
                    "select": [
                        "ID", "CONTACT_ID", "DATE_CREATE", "STAGE_ID",
                        "UF_CRM_1738231898", "UF_CRM_1738232000",
                    ],
                    "start": start,
                }, retries=1)
            except Exception as e:
                # Если уже собрали часть данных — возвращаем что есть
                if all_deals:
                    print(f"[api] Частичный сбор: {len(all_deals)} сделок (прервано: {e})")
                raise
            batch = result.get("result", [])
            if not batch:
                break
            all_deals.extend(batch)
            nxt = result.get("next")
            if not nxt or nxt <= start:
                break
            start = nxt
        return all_deals

    def _build_visits(self) -> None:
        """Построить структуру: doctor_id -> contact_id -> [visit dates]."""
        visits: dict[int, dict[int, list[date]]] = defaultdict(lambda: defaultdict(list))
        for d in self.deals:
            doctor_raw = d.get("UF_CRM_1738231898")
            contact_raw = d.get("CONTACT_ID")
            date_raw = d.get("DATE_CREATE")
            if not doctor_raw or not contact_raw or not date_raw:
                continue
            try:
                doctor_id = int(doctor_raw)
                contact_id = int(contact_raw)
            except (ValueError, TypeError):
                continue
            if doctor_id not in ACTIVE_IDS:
                continue
            try:
                visit_date = datetime.fromisoformat(date_raw).date()
            except (ValueError, TypeError):
                # fallback: parse YYYY-MM-DD
                try:
                    visit_date = datetime.strptime(str(date_raw)[:10], "%Y-%m-%d").date()
                except ValueError:
                    continue
            if visit_date < DATA_START:
                continue
            visits[doctor_id][contact_id].append(visit_date)
        # Сортируем даты каждого пациента
        for doctor_id in visits:
            for contact_id in visits[doctor_id]:
                visits[doctor_id][contact_id] = sorted(visits[doctor_id][contact_id])
        self.visits = dict(visits)

    # -------------------------------------------------------------------------
    # Retention calculation
    # -------------------------------------------------------------------------

    def _primary_visits_in_window(
        self, window_start: date, window_end: date
    ) -> dict[int, dict[int, date]]:
        """
        Для заданного временного окна найти ПЕРВЫЙ визит каждого (врач, пациент).
        Возвращает: doctor_id -> {contact_id: first_visit_date}
        """
        result: dict[int, dict[int, date]] = defaultdict(dict)
        for doctor_id, contacts in self.visits.items():
            for contact_id, dates in contacts.items():
                in_window = [d for d in dates if window_start <= d < window_end]
                if in_window:
                    result[doctor_id][contact_id] = min(in_window)
        return dict(result)

    def _check_return(
        self, contact_id: int, from_date: date, window_months: int,
        same_doctor_id: int | None = None,
    ) -> tuple[bool, bool, int | None]:
        """
        Проверить, вернулся ли пациент в пределах N месяцев.
        Returns: (returned_same_doctor, returned_any, next_doctor_id)
        next_doctor_id — врач хронологически первого повторного визита (для матрицы перетекания).
        """
        end_date = add_months(from_date, window_months)
        end_date = min(end_date, self.report_date)

        returned_same = False
        returned_any = False
        next_doctor = None
        earliest_return_date: date | None = None

        for doctor_id, contacts in self.visits.items():
            for d in contacts.get(contact_id, []):
                if from_date < d <= end_date:
                    returned_any = True
                    if same_doctor_id is not None and doctor_id == same_doctor_id:
                        returned_same = True
                    # Хронологически самый ранний возврат
                    if earliest_return_date is None or d < earliest_return_date:
                        earliest_return_date = d
                        next_doctor = doctor_id

        return returned_same, returned_any, next_doctor

    def calculate_window(self, months: int) -> dict:
        """
        Рассчитать метрики возвращаемости для окна N месяцев.
        Возвращает словарь с Rate A, Rate B, матрицей перетекания.
        """
        window_end = self.report_date
        window_start = add_months(self.report_date, -months)

        primaries = self._primary_visits_in_window(window_start, window_end)

        # Rate A: return to same doctor
        # Rate B: return to any doctor
        # Transfer D: doctor X -> doctor Y
        doctor_stats: dict[int, dict] = defaultdict(lambda: {
            "primary_count": 0,       # всего первичных
            "returned_same": 0,       # вернулись к тому же
            "returned_any": 0,        # вернулись в клинику
            "rate_a": None,           # Return Rate A
            "rate_b": None,           # Return Rate B
            "primary_patients": [],   # список contact_id
        })

        transfer: dict[int, dict[int, int]] = defaultdict(lambda: defaultdict(int))

        for doctor_id, contacts in primaries.items():
            for contact_id, first_date in contacts.items():
                doctor_stats[doctor_id]["primary_count"] += 1
                doctor_stats[doctor_id]["primary_patients"].append(contact_id)

                same, any_doc, next_doc = self._check_return(
                    contact_id, first_date, months, same_doctor_id=doctor_id
                )
                if same:
                    doctor_stats[doctor_id]["returned_same"] += 1
                if any_doc:
                    doctor_stats[doctor_id]["returned_any"] += 1
                if next_doc is not None and next_doc != doctor_id:
                    transfer[doctor_id][next_doc] += 1

        # Рассчитать ставки
        for doctor_id, stats in doctor_stats.items():
            n = stats["primary_count"]
            stats["rate_a"] = stats["returned_same"] / n if n > 0 else 0.0
            stats["rate_b"] = stats["returned_any"] / n if n > 0 else 0.0

        # Среднее по клинике (только врачи с >= MIN_PRIMARY_PATIENTS)
        significant_rates_a = [
            s["rate_a"] for s in doctor_stats.values()
            if s["primary_count"] >= MIN_PRIMARY_PATIENTS
        ]
        significant_rates_b = [
            s["rate_b"] for s in doctor_stats.values()
            if s["primary_count"] >= MIN_PRIMARY_PATIENTS
        ]
        avg_a = sum(significant_rates_a) / len(significant_rates_a) if significant_rates_a else 0.0
        avg_b = sum(significant_rates_b) / len(significant_rates_b) if significant_rates_b else 0.0

        return {
            "window_months": months,
            "window_start": window_start.isoformat(),
            "window_end": window_end.isoformat(),
            "doctor_stats": dict(doctor_stats),
            "avg_rate_a": avg_a,
            "avg_rate_b": avg_b,
            "transfer_matrix": dict(transfer),
        }

    def calculate_all_windows(self) -> dict[int, dict]:
        """Рассчитать все 4 окна."""
        results = {}
        for m in WINDOWS:
            results[m] = self.calculate_window(m)
        return results

    def calculate_trends(self, months: int, num_periods: int = 6) -> dict[int, list[float]]:
        """
        Рассчитать тренды: Rate A для последних N периодов окна данной длины.
        """
        trends: dict[int, list[float]] = defaultdict(list)
        for p in range(num_periods):
            # period_end = report_date - p * months
            period_end = add_months(self.report_date, -p * months)
            period_start = add_months(period_end, -months)
            primaries = self._primary_visits_in_window(period_start, period_end)

            period_rates: dict[int, float | None] = {}
            for doctor_id, contacts in primaries.items():
                returned = 0
                total = len(contacts)
                for contact_id, first_date in contacts.items():
                    same, _, _ = self._check_return(contact_id, first_date, months, same_doctor_id=doctor_id)
                    if same:
                        returned += 1
                period_rates[doctor_id] = (returned / total * 100) if total > 0 else None

            for doctor_id in ACTIVE_IDS:
                trends[doctor_id].append(period_rates.get(doctor_id))

        return dict(trends)

    # -------------------------------------------------------------------------
    # Forecast
    # -------------------------------------------------------------------------

    def calculate_forecast(self, window_3m: dict | None = None) -> dict:
        """Прогноз загрузки на следующий месяц."""
        # Среднее число первичных пациентов в месяц за последние 6 мес
        forecast_start = add_months(self.report_date, -6)
        primaries = self._primary_visits_in_window(forecast_start, self.report_date)
        monthly_avg: dict[int, float] = {}
        for doctor_id, contacts in primaries.items():
            monthly_avg[doctor_id] = len(contacts) / 6.0

        # Историческая возвращаемость (3 мес)
        w3 = window_3m if window_3m is not None else self.calculate_window(3)

        # Средний чек врача за последние 6 мес
        avg_check: dict[int, float] = defaultdict(float)
        deal_count: dict[int, int] = defaultdict(int)
        for d in self.deals:
            amount_raw = d.get("UF_CRM_1738232000")
            doctor_raw = d.get("UF_CRM_1738231898")
            date_raw = d.get("DATE_CREATE")
            if not amount_raw or not doctor_raw or not date_raw:
                continue
            try:
                doctor_id = int(doctor_raw)
                amount = float(amount_raw)
                deal_date = datetime.fromisoformat(str(date_raw)).date()
            except (ValueError, TypeError):
                continue
            if deal_date >= forecast_start and doctor_id in ACTIVE_IDS:
                avg_check[doctor_id] += amount
                deal_count[doctor_id] += 1
        for d_id in avg_check:
            if deal_count[d_id] > 0:
                avg_check[d_id] = avg_check[d_id] / deal_count[d_id]

        # Прогноз
        forecast: dict[int, dict] = {}
        for doctor_id in ACTIVE_IDS:
            rate_a_3m = (w3.get("doctor_stats", {}).get(doctor_id, {}).get("rate_a") or 0)
            prim_new = monthly_avg.get(doctor_id, 0)
            prim_return = prim_new * rate_a_3m
            total_expected = prim_new + prim_return
            check = avg_check.get(doctor_id, 0)
            forecast[doctor_id] = {
                "expected_new": round(prim_new, 1),
                "expected_return": round(prim_return, 1),
                "expected_total": round(total_expected, 1),
                "avg_check": round(check, 0),
                "expected_revenue": round(total_expected * check, 0),
            }
        return forecast

    # -------------------------------------------------------------------------
    # XLSX Report
    # -------------------------------------------------------------------------

    def generate_xlsx(self, window_results: dict[int, dict], trends: dict,
                      forecast: dict[int, dict]) -> str:
        """Сформировать XLSX-отчёт, вернуть путь к файлу."""
        os.makedirs(self.output_dir, exist_ok=True)
        filename = f"retention-{self.report_date.isoformat()}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = openpyxl.Workbook()

        self._sheet_summary(wb, window_results)
        self._sheet_rating(wb, window_results)
        self._sheet_trends(wb, trends)
        self._sheet_transfer(wb, window_results)
        self._sheet_forecast(wb, forecast)

        wb.save(filepath)
        return filepath

    def _style_header_row(self, ws, headers: list[str], row: int = 1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def _style_cell(self, ws, row: int, col: int, value, fmt: str | None = None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if isinstance(value, (int, float)) else "left")
        return cell

    def _sheet_summary(self, wb, results: dict[int, dict]):
        ws = wb.active
        ws.title = "Сводка"

        headers = ["Врач"]
        for m in WINDOWS:
            headers.append(f"Rate A {m}мес")
            headers.append(f"Rate B {m}мес")
            headers.append(f"Первичных {m}мес")
        self._style_header_row(ws, headers)

        row = 2
        for doctor_id in ACTIVE_IDS:
            name = DOCTOR_MAP.get(doctor_id, str(doctor_id))
            col = 1
            self._style_cell(ws, row, col, name)
            col += 1
            for m in WINDOWS:
                stats = results.get(m, {}).get("doctor_stats", {}).get(doctor_id, {})
                rate_a = stats.get("rate_a")
                rate_b = stats.get("rate_b")
                prim = stats.get("primary_count", 0)
                cell_a = self._style_cell(ws, row, col, format_rate(rate_a) if rate_a is not None else "—")
                col += 1
                cell_b = self._style_cell(ws, row, col, format_rate(rate_b) if rate_b is not None else "—")
                col += 1
                cell_p = self._style_cell(ws, row, col, prim)
                col += 1
                # Color coding
                avg_a = results.get(m, {}).get("avg_rate_a", 0)
                if rate_a is not None and prim >= MIN_PRIMARY_PATIENTS:
                    if rate_a >= avg_a:
                        cell_a.fill = GREEN_FILL
                    else:
                        cell_a.fill = RED_FILL
                elif prim < MIN_PRIMARY_PATIENTS and prim > 0:
                    cell_a.fill = YELLOW_FILL
            row += 1

        # Averages row
        self._style_cell(ws, row, 1, "Среднее по клинике")
        col = 2
        for m in WINDOWS:
            avg_a = results.get(m, {}).get("avg_rate_a", 0)
            avg_b = results.get(m, {}).get("avg_rate_b", 0)
            cell_a = self._style_cell(ws, row, col, format_rate(avg_a))
            cell_a.fill = LIGHT_BLUE_FILL
            col += 1
            cell_b = self._style_cell(ws, row, col, format_rate(avg_b))
            cell_b.fill = LIGHT_BLUE_FILL
            col += 1
            self._style_cell(ws, row, col, "—")
            col += 1

        ws.column_dimensions["A"].width = 30
        for i in range(2, 1 + len(WINDOWS) * 3 + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

    def _sheet_rating(self, wb, results: dict[int, dict]):
        ws = wb.create_sheet("Рейтинг")

        headers = ["Рейтинг", "Врач"]
        for m in WINDOWS:
            headers.append(f"Rate A {m}мес")
            headers.append(f"Откл. {m}мес")
        headers.append("Первичных (3мес)")
        self._style_header_row(ws, headers)

        # Сортируем по Rate A 3 мес
        def sort_key(d_id):
            stats = results.get(3, {}).get("doctor_stats", {}).get(d_id, {})
            return -(stats.get("rate_a") or 0)
        sorted_doctors = sorted(ACTIVE_IDS, key=sort_key)

        row = 2
        for rank, doctor_id in enumerate(sorted_doctors, 1):
            name = DOCTOR_MAP.get(doctor_id, str(doctor_id))
            col = 1
            self._style_cell(ws, row, col, rank)
            col += 1
            self._style_cell(ws, row, col, name)
            col += 1
            for m in WINDOWS:
                stats = results.get(m, {}).get("doctor_stats", {}).get(doctor_id, {})
                rate_a = stats.get("rate_a")
                avg_a = results.get(m, {}).get("avg_rate_a", 0)
                dev = (rate_a - avg_a) * 100 if rate_a is not None else None

                cell_r = self._style_cell(ws, row, col, format_rate(rate_a) if rate_a is not None else "—")
                col += 1
                cell_d = self._style_cell(ws, row, col, f"{dev:+.1f}%" if dev is not None else "—")
                col += 1

                prim = stats.get("primary_count", 0)
                if rate_a is not None and prim >= MIN_PRIMARY_PATIENTS:
                    if rate_a >= avg_a:
                        cell_r.fill = GREEN_FILL
                    else:
                        cell_r.fill = RED_FILL
                elif prim > 0:
                    cell_r.fill = YELLOW_FILL

            # Primary count
            p3 = results.get(3, {}).get("doctor_stats", {}).get(doctor_id, {}).get("primary_count", 0)
            self._style_cell(ws, row, 1 + 2 + len(WINDOWS) * 2, p3)
            row += 1

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        for i in range(3, 3 + len(WINDOWS) * 2 + 1):
            ws.column_dimensions[get_column_letter(i)].width = 14

    def _sheet_trends(self, wb, trends: dict[int, list[float | None]]):
        ws = wb.create_sheet("Тренды")
        num_periods = len(next(iter(trends.values()), []))
        headers = ["Врач"]
        for i in range(num_periods, 0, -1):
            headers.append(f"Период -{i} (3мес)")
        self._style_header_row(ws, headers)

        row = 2
        for doctor_id in ACTIVE_IDS:
            name = DOCTOR_MAP.get(doctor_id, str(doctor_id))
            self._style_cell(ws, row, 1, name)
            col = 2
            vals = trends.get(doctor_id, [])
            for v in reversed(vals):
                val_str = f"{v:.1f}%" if v is not None else "—"
                cell = self._style_cell(ws, row, col, val_str)
                if v is not None and vals.count(None) < len(vals):
                    # Color based on direction vs average
                    pass
                col += 1
            row += 1

        ws.column_dimensions["A"].width = 30
        for i in range(2, 2 + num_periods):
            ws.column_dimensions[get_column_letter(i)].width = 16

    def _sheet_transfer(self, wb, results: dict[int, dict]):
        ws = wb.create_sheet("Перетекание")
        # Use 12-month window for transfer matrix
        transfer = results.get(12, {}).get("transfer_matrix", {})

        sorted_ids = sorted(ACTIVE_IDS)
        headers = ["От врача \\ К врачу"] + [DOCTOR_MAP.get(d, str(d)) for d in sorted_ids]
        self._style_header_row(ws, headers)

        row = 2
        for from_id in sorted_ids:
            self._style_cell(ws, row, 1, DOCTOR_MAP.get(from_id, str(from_id)))
            col = 2
            from_row = transfer.get(from_id, {})
            for to_id in sorted_ids:
                count = from_row.get(to_id, 0)
                cell = self._style_cell(ws, row, col, count if count > 0 else "")
                if count > 0:
                    cell.fill = YELLOW_FILL
                col += 1
            row += 1

        ws.column_dimensions["A"].width = 30
        for i in range(2, 2 + len(sorted_ids)):
            ws.column_dimensions[get_column_letter(i)].width = 14

    def _sheet_forecast(self, wb, forecast: dict[int, dict]):
        ws = wb.create_sheet("Прогноз")
        headers = ["Врач", "Ожид. новых", "Ожид. возвратов", "Всего визитов",
                    "Средний чек", "Ожид. выручка"]
        self._style_header_row(ws, headers)

        row = 2
        total_new = 0.0
        total_ret = 0.0
        total_rev = 0.0
        for doctor_id in ACTIVE_IDS:
            f = forecast.get(doctor_id, {})
            name = DOCTOR_MAP.get(doctor_id, str(doctor_id))
            self._style_cell(ws, row, 1, name)
            self._style_cell(ws, row, 2, f.get("expected_new", 0))
            self._style_cell(ws, row, 3, f.get("expected_return", 0))
            self._style_cell(ws, row, 4, f.get("expected_total", 0))
            self._style_cell(ws, row, 5, f"{f.get('avg_check', 0):,.0f} ₽")
            self._style_cell(ws, row, 6, f"{f.get('expected_revenue', 0):,.0f} ₽")
            total_new += f.get("expected_new", 0)
            total_ret += f.get("expected_return", 0)
            total_rev += f.get("expected_revenue", 0)
            row += 1

        # Total row
        self._style_cell(ws, row, 1, "ИТОГО")
        cell = self._style_cell(ws, row, 2, round(total_new, 1))
        cell.fill = LIGHT_BLUE_FILL
        cell = self._style_cell(ws, row, 3, round(total_ret, 1))
        cell.fill = LIGHT_BLUE_FILL
        cell = self._style_cell(ws, row, 4, round(total_new + total_ret, 1))
        cell.fill = LIGHT_BLUE_FILL
        self._style_cell(ws, row, 5, "—")
        cell = self._style_cell(ws, row, 6, f"{total_rev:,.0f} ₽")
        cell.fill = LIGHT_BLUE_FILL

        ws.column_dimensions["A"].width = 30
        for c in "BCDEF":
            ws.column_dimensions[c].width = 18

    # -------------------------------------------------------------------------
    # Markdown summary
    # -------------------------------------------------------------------------

    def generate_markdown(self, window_results: dict[int, dict], trends: dict,
                          forecast: dict[int, dict]) -> str:
        """Сгенерировать Markdown-сводку."""
        lines = []
        lines.append(f"# Возвращаемость пациентов — {self.report_date.isoformat()}")
        lines.append("")

        # Baseline
        lines.append("## Бенчмарк (среднее по клинике)")
        for m in WINDOWS:
            avg_a = window_results.get(m, {}).get("avg_rate_a", 0)
            avg_b = window_results.get(m, {}).get("avg_rate_b", 0)
            lines.append(f"- **{m} мес:** Rate A = {format_rate(avg_a)}, Rate B = {format_rate(avg_b)}")
        lines.append("")

        # Top-3
        lines.append("## Топ-3 по возвращаемости (Rate A, 3 мес)")
        w3 = window_results.get(3, {})
        stats3 = w3.get("doctor_stats", {})
        sorted_by_a = sorted(
            [(d_id, s.get("rate_a", 0), s.get("primary_count", 0))
             for d_id, s in stats3.items() if s.get("primary_count", 0) >= MIN_PRIMARY_PATIENTS],
            key=lambda x: -x[1],
        )
        for i, (d_id, rate, count) in enumerate(sorted_by_a[:3], 1):
            name = DOCTOR_MAP.get(d_id, str(d_id))
            lines.append(f"{i}. **{name}** — {format_rate(rate)} ({count} первичных)")
        lines.append("")

        # Underperformers (alert)
        lines.append("## Аутсайдеры и алерты")
        avg_a_3 = w3.get("avg_rate_a", 0)
        alerts = []
        for doctor_id in ACTIVE_IDS:
            s = stats3.get(doctor_id, {})
            rate = s.get("rate_a")
            count = s.get("primary_count", 0)
            if rate is not None and count >= MIN_PRIMARY_PATIENTS and rate < avg_a_3:
                name = DOCTOR_MAP.get(doctor_id, str(doctor_id))
                deviation = (rate - avg_a_3) * 100
                alerts.append((name, rate, count, deviation))
        if alerts:
            alerts.sort(key=lambda x: x[1])  # sort by rate ascending
            for name, rate, count, dev in alerts:
                lines.append(f"- [!] **{name}**: {format_rate(rate)} при среднем {format_rate(avg_a_3)} "
                             f"(отклонение {dev:+.1f}%, {count} первичных)")
        else:
            lines.append("Все врачи на уровне или выше среднего.")
        lines.append("")

        # Transfer highlights
        lines.append("## Перетекание пациентов (12 мес)")
        transfer = window_results.get(12, {}).get("transfer_matrix", {})
        top_transfers = []
        for from_id, to_dict in transfer.items():
            from_name = DOCTOR_MAP.get(from_id, str(from_id))
            for to_id, count in to_dict.items():
                to_name = DOCTOR_MAP.get(to_id, str(to_id))
                if count >= 2:
                    top_transfers.append((count, from_name, to_name))
        top_transfers.sort(key=lambda x: -x[0])
        if top_transfers:
            for count, from_name, to_name in top_transfers[:5]:
                lines.append(f"- {from_name} → {to_name}: **{count} пациентов**")
        else:
            lines.append("Значимых перетеканий не выявлено.")
        lines.append("")

        # Forecast
        lines.append("## Прогноз загрузки на следующий месяц")
        total_new = sum(f.get("expected_new", 0) for f in forecast.values())
        total_ret = sum(f.get("expected_return", 0) for f in forecast.values())
        total_rev = sum(f.get("expected_revenue", 0) for f in forecast.values())
        lines.append(f"- Ожидаемое число новых первичных: **{total_new:.0f}**")
        lines.append(f"- Ожидаемое число возвратов: **{total_ret:.0f}**")
        lines.append(f"- Всего визитов: **{total_new + total_ret:.0f}**")
        lines.append(f"- Ожидаемая выручка: **{total_rev:,.0f} ₽**")
        lines.append("")

        return "\n".join(lines)

    # -------------------------------------------------------------------------
    # Weekly quick check
    # -------------------------------------------------------------------------

    def weekly_check(self) -> str | None:
        """
        Быстрая проверка Rate A за 3 месяца.
        Возвращает строку с алертами, или None если всё в норме.
        """
        w3 = self.calculate_window(3)
        avg_a = w3["avg_rate_a"]
        stats = w3["doctor_stats"]

        alerts = []
        for doctor_id in ACTIVE_IDS:
            s = stats.get(doctor_id, {})
            rate = s.get("rate_a")
            count = s.get("primary_count", 0)
            if rate is not None and count >= MIN_PRIMARY_PATIENTS and rate < avg_a:
                name = DOCTOR_MAP.get(doctor_id, str(doctor_id))
                deviation = (rate - avg_a) * 100
                alerts.append(
                    f"[!] Врач {name}: возвращаемость {format_rate(rate)} "
                    f"при среднем по клинике {format_rate(avg_a)}. "
                    f"Отклонение {deviation:+.1f}%."
                )
        if alerts:
            return "\n".join(alerts)
        return None

    # -------------------------------------------------------------------------
    # Run modes
    # -------------------------------------------------------------------------

    def run_monthly(self, force_refresh: bool = False) -> str:
        """Полный месячный отчёт. Возвращает путь к XLSX."""
        print(f"[monthly] Расчёт возвращаемости на {self.report_date.isoformat()}")
        self.load_data(force_refresh=force_refresh)
        if not self.deals:
            print("[monthly] Нет данных для расчёта.")
            return ""

        results = self.calculate_all_windows()
        trends = self.calculate_trends(3, num_periods=6)
        forecast = self.calculate_forecast(window_3m=results.get(3))

        # XLSX
        xlsx_path = self.generate_xlsx(results, trends, forecast)
        print(f"[monthly] Отчёт сохранён: {xlsx_path}")

        # Markdown
        md = self.generate_markdown(results, trends, forecast)
        print("\n" + md)

        return xlsx_path

    def run_weekly(self, force_refresh: bool = False) -> int:
        """
        Еженедельная быстрая проверка.
        Выводит алерты в stdout. Возвращает число алертов (0 = всё в норме).
        """
        self.load_data(force_refresh=force_refresh, silent=True)
        if not self.deals:
            return 0
        alert = self.weekly_check()
        if alert:
            print(alert)
            return alert.count("[!]")
        return 0


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="ARclinic Doctor Retention Agent")
    parser.add_argument("mode", choices=["monthly", "weekly"], help="Режим работы")
    parser.add_argument("--date", type=str, default=None,
                        help="Дата отчёта (YYYY-MM-DD), по умолчанию сегодня")
    parser.add_argument("--force-refresh", action="store_true",
                        help="Принудительно перезапросить данные из API")
    parser.add_argument("--output-dir", type=str, default="reports",
                        help="Директория для отчётов (по умолчанию reports)")
    args = parser.parse_args()

    report_date = date.today()
    if args.date:
        try:
            report_date = datetime.strptime(args.date, "%Y-%m-%d").date()
        except ValueError:
            print(f"Неверный формат даты: {args.date}. Используйте YYYY-MM-DD.", file=sys.stderr)
            sys.exit(1)

    agent = DoctorRetentionAgent(report_date=report_date, output_dir=args.output_dir)

    if args.mode == "monthly":
        agent.run_monthly(force_refresh=args.force_refresh)
    elif args.mode == "weekly":
        agent.run_weekly(force_refresh=args.force_refresh)


if __name__ == "__main__":
    main()
