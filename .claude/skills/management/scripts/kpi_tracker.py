import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date
import os


def generate_kpi_report(kpi_data, output_dir="reports"):
    os.makedirs(output_dir, exist_ok=True)
    filename = f"kpi-{date.today().isoformat()}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Sheet 1 - Individual KPIs
    ws1 = wb.active
    ws1.title = "KPI врачей"

    ind_headers = ["Врач", "Загрузка %", "Конверсия %", "Средний чек", "Новые пациенты", "Донаты", "Рейтинг"]
    for col, h in enumerate(ind_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    doctors = kpi_data.get("doctors", [])
    for i, d in enumerate(doctors, 2):
        ws1.cell(row=i, column=1, value=d.get("name", "")).border = thin_border
        for col, key in enumerate(["load_pct", "conversion_pct", "avg_check", "new_patients", "donations", "rating"], 2):
            cell = ws1.cell(row=i, column=col, value=d.get(key, "нет данных"))
            cell.border = thin_border

    ws1.column_dimensions['A'].width = 25
    for c in 'BCDEFG':
        ws1.column_dimensions[c].width = 16

    # Sheet 2 - Team KPIs
    ws2 = wb.create_sheet("Командные KPI")
    team_headers = ["Метрика", "Текущее", "План", "Выполнение %", "Динамика"]

    for col, h in enumerate(team_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    team_metrics = kpi_data.get("team", [])
    for i, m in enumerate(team_metrics, 2):
        ws2.cell(row=i, column=1, value=m.get("metric", "")).border = thin_border
        ws2.cell(row=i, column=2, value=m.get("current", "нет данных")).border = thin_border
        ws2.cell(row=i, column=3, value=m.get("plan", "нет данных")).border = thin_border
        pct = m.get("progress_pct")
        cell = ws2.cell(row=i, column=4, value=pct if pct is not None else "нет данных")
        cell.border = thin_border
        if pct is not None:
            if pct >= 90:
                cell.fill = green_fill
            elif pct >= 70:
                cell.fill = yellow_fill
            else:
                cell.fill = red_fill
        ws2.cell(row=i, column=5, value=m.get("trend", "нет данных")).border = thin_border

    ws2.column_dimensions['A'].width = 25
    for c in 'BCDE':
        ws2.column_dimensions[c].width = 16

    wb.save(filepath)
    return filepath
